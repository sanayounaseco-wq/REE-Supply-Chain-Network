"""
STEP 5 — Panel Regressions (Models 1 & 2) — Corrected Specifications
══════════════════════════════════════════════════════════════════════
Model 1 (Vulnerability):
  DV: china_import_share_L2
  IV: alliance_diversity_idx, msp_member, n_suppliers_L2, log_gdp
  Logic: What reduces a country's dependence on China for midstream REE?

Model 2 (Bargaining Power):
  DV: dependent_importers_25pct
  IV: betweenness_L2, betweenness_L1, export_restriction_binary, log_gdp
  Logic: Does Layer 2 bottleneck centrality predict how many countries
         depend on you, controlling for mining production and policy?

Sub-specifications test H2 directly:
  2b adds log_mining to test: does centrality predict power OVER AND ABOVE raw production?
  2c compares layer centrality coefficients (L1 vs L2 — H2 test)

HOW TO RUN:
    pip install linearmodels statsmodels openpyxl
    python Step5_Regressions.py
"""

import pandas as pd
import numpy as np
import statsmodels.formula.api as smf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os, warnings
warnings.filterwarnings("ignore")

DATA_DIR    = "/Users/sanayounas/Claude/Projects/Netwrok/data"
PROJECT_DIR = "/Users/sanayounas/Claude/Projects/Netwrok"

# ── LOAD & PREPARE ─────────────────────────────────────────────
print("Loading master panel...")
df = pd.read_csv(os.path.join(DATA_DIR, "04_master_panel.csv"))
print(f"  {len(df):,} rows | {df['iso3'].nunique()} countries | {df['year'].nunique()} years")

# Diagnostic: print available columns so we know what WDI data landed
print(f"  Columns in master panel: {sorted(df.columns.tolist())}\n")

# Safe log transform — skips column if it doesn't exist
def safe_log(df, src_col, dest_col):
    if src_col in df.columns:
        df[dest_col] = np.log(df[src_col].replace(0, np.nan))
    else:
        print(f"  ⚠ Column '{src_col}' not found — '{dest_col}' will be omitted from models")
        df[dest_col] = np.nan

safe_log(df, "gdp_current_usd",      "log_gdp")
safe_log(df, "population",           "log_pop")
safe_log(df, "mining_production_REO","log_mining")
safe_log(df, "out_strength_L2",      "log_out_L2")
safe_log(df, "in_strength_L2",       "log_in_L2")

# Winsorise HHI and betweenness at 99th percentile
for col in ["supplier_hhi_L2", "betweenness_L2", "betweenness_L1", "betweenness_L3"]:
    p99 = df[col].quantile(0.99)
    df[f"{col}_w"] = df[col].clip(upper=p99)

# Restrict to country-years with actual Layer 2 trade
has_L2 = df["in_strength_L2"].notna() & (df["in_strength_L2"] > 0)
has_L2_exp = df["out_strength_L2"].notna() & (df["out_strength_L2"] > 0)
df_imp = df[has_L2].copy()      # importer perspective (Model 1)
df_exp = df[has_L2_exp].copy()  # exporter perspective (Model 2)

print(f"  Model 1 sample (importers with L2 trade): {len(df_imp):,} rows | {df_imp['iso3'].nunique()} countries")
print(f"  Model 2 sample (exporters with L2 trade): {len(df_exp):,} rows | {df_exp['iso3'].nunique()} countries")

# ── HELPER FUNCTIONS ───────────────────────────────────────────
def run_ols_fe(data, formula, label=""):
    clean = data.copy()
    # Extract variable names from formula
    dep = formula.split("~")[0].strip()
    clean = clean.dropna(subset=[dep])
    try:
        model = smf.ols(formula, data=clean).fit(cov_type="HC1")
        n_obs = int(model.nobs)
        print(f"  {label}: N={n_obs}, R²={model.rsquared:.4f}")
        return model
    except Exception as e:
        print(f"  ERROR [{label}]: {e}")
        return None

def extract_coef_table(model, var_map, title=""):
    """var_map: {internal_name: display_name}"""
    if model is None:
        return pd.DataFrame()
    rows = []
    for var, label in var_map.items():
        if var in model.params.index:
            p = model.pvalues[var]
            rows.append({
                "Variable": label,
                "Coef.": round(model.params[var], 4),
                "Std.Err.": round(model.bse[var], 4),
                "t": round(model.tvalues[var], 3),
                "p-value": round(p, 4),
                "": "***" if p<0.01 else "**" if p<0.05 else "*" if p<0.10 else "",
            })
    rows += [
        {"Variable": "Country FE",       "Coef.": "Yes", "Std.Err.": "", "t": "", "p-value": "", "": ""},
        {"Variable": "Year FE",           "Coef.": "Yes", "Std.Err.": "", "t": "", "p-value": "", "": ""},
        {"Variable": "N",                 "Coef.": int(model.nobs), "Std.Err.": "", "t": "", "p-value": "", "": ""},
        {"Variable": "R² (within)",       "Coef.": round(model.rsquared, 4), "Std.Err.": "", "t": "", "p-value": "", "": ""},
        {"Variable": "Adj. R²",           "Coef.": round(model.rsquared_adj, 4), "Std.Err.": "", "t": "", "p-value": "", "": ""},
    ]
    return pd.DataFrame(rows)

# ══════════════════════════════════════════════════════════════
# MODEL 1 — SUPPLY VULNERABILITY (importer-side)
# DV: china_import_share_L2  (direct, not composite)
# Why: measures China dependence directly; avoids mechanical R²=1 problem
# ══════════════════════════════════════════════════════════════
print("\n═══ MODEL 1: Supply Vulnerability (DV = China import share, Layer 2) ═══")
print("H3: Higher supplier concentration → higher China dependence")
print("H4: Allied diversification (MSP) → lower China dependence\n")

# M1a — Main specification
m1a_vars = {
    "alliance_diversity_idx": "Alliance diversity index  [H4: expect −]",
    "msp_member":             "MSP member (post-2022)    [H4: expect −]",
    "n_suppliers_L2":         "Number of suppliers (L2)  [H3: expect −]",
    "supplier_hhi_L2_w":      "Supplier HHI (winsorised) [H3: expect +]",
    "log_gdp":                "log(GDP)",
}
m1a_formula = ("china_import_share_L2 ~ "
               "alliance_diversity_idx + msp_member + n_suppliers_L2 + "
               "supplier_hhi_L2_w + log_gdp + C(iso3) + C(year)")
m1a = run_ols_fe(df_imp, m1a_formula, "M1a Main")
res_m1a = extract_coef_table(m1a, m1a_vars)

# M1b — Add betweenness of country's top supplier (network channel)
m1b_vars = {**m1a_vars,
    "betweenness_L2_w": "Betweenness L2 (own network position)",
}
m1b_formula = ("china_import_share_L2 ~ "
               "alliance_diversity_idx + msp_member + n_suppliers_L2 + "
               "supplier_hhi_L2_w + betweenness_L2_w + log_gdp + C(iso3) + C(year)")
m1b = run_ols_fe(df_imp, m1b_formula, "M1b + own betweenness")
res_m1b = extract_coef_table(m1b, m1b_vars)

# M1c — Alternative DV: supply_vulnerability_idx with DIFFERENT regressors (no components)
m1c_vars = {
    "betweenness_L2_w":       "Own betweenness L2 (network position)",
    "msp_member":             "MSP member (post-2022)   [H4: expect −]",
    "alliance_diversity_idx": "Alliance diversity index [H4: expect −]",
    "export_restriction_binary": "Faces export restriction (own)",
    "log_gdp":                "log(GDP)",
}
m1c_formula = ("supply_vulnerability_idx ~ "
               "betweenness_L2_w + msp_member + alliance_diversity_idx + "
               "export_restriction_binary + log_gdp + C(iso3) + C(year)")
m1c = run_ols_fe(df_imp, m1c_formula, "M1c SVI ~ network vars only")
res_m1c = extract_coef_table(m1c, m1c_vars)

# Print M1a results
if m1a:
    print("\nM1a Results:")
    for _, row in res_m1a.iterrows():
        if row["Variable"] not in ("Country FE","Year FE","N","R² (within)","Adj. R²"):
            print(f"  {row['Variable']:<45} {str(row['Coef.']):>10}  {row['']:>4}")

# ══════════════════════════════════════════════════════════════
# MODEL 2 — BARGAINING POWER (exporter-side)
# DV: dependent_importers_25pct
# Key test: does betweenness_L2 > betweenness_L1 (H2)?
# ══════════════════════════════════════════════════════════════
print("\n═══ MODEL 2: Bargaining Power (DV = Dependent importers at 25%) ═══")
print("H1: Higher betweenness centrality → more dependent importers")
print("H2: L2 betweenness coefficient > L1 betweenness coefficient\n")

# M2a — Core specification: betweenness L2 only (max N)
has_pop = df["log_pop"].notna().sum() > 100
pop_term = "+ log_pop" if has_pop else ""
pop_var  = {"log_pop": "log(Population)"} if has_pop else {}

m2a_vars = {
    "betweenness_L2_w":          "Betweenness centrality L2 (midstream)  [H1: expect +]",
    "export_restriction_binary":  "Export restriction binary (OECD)",
    "log_gdp":                    "log(GDP)",
    **pop_var,
}
m2a_formula = ("dependent_importers_25pct ~ "
               f"betweenness_L2_w + export_restriction_binary + log_gdp {pop_term} + C(iso3) + C(year)")
m2a = run_ols_fe(df_exp, m2a_formula, "M2a L2 betweenness only")
res_m2a = extract_coef_table(m2a, m2a_vars)

# M2b — Add L1 betweenness for layer comparison (H2 direct test)
m2b_vars = {
    "betweenness_L2_w":          "Betweenness L2 — midstream  [H2: expect β₂ > β₁]",
    "betweenness_L1_w":          "Betweenness L1 — upstream   [H2: comparison]",
    "export_restriction_binary":  "Export restriction binary",
    "log_gdp":                    "log(GDP)",
    **pop_var,
}
m2b_formula = ("dependent_importers_25pct ~ "
               f"betweenness_L2_w + betweenness_L1_w + export_restriction_binary + "
               f"log_gdp {pop_term} + C(iso3) + C(year)")
m2b = run_ols_fe(df_exp, m2b_formula, "M2b L2+L1 betweenness (H2 test)")
res_m2b = extract_coef_table(m2b, m2b_vars)

# M2c — Add log_mining to test H2: centrality > production
m2c_vars = {
    "betweenness_L2_w":          "Betweenness L2 — midstream  [H2: expect β > mining]",
    "betweenness_L1_w":          "Betweenness L1 — upstream",
    "log_mining":                 "log(Mining production, REO)  [H2: expect < betweenness]",
    "export_restriction_binary":  "Export restriction binary",
    "log_gdp":                    "log(GDP)",
}
m2c_formula = ("dependent_importers_25pct ~ "
               "betweenness_L2_w + betweenness_L1_w + log_mining + "
               "export_restriction_binary + log_gdp + C(iso3) + C(year)")
m2c = run_ols_fe(df_exp, m2c_formula, "M2c + mining production (H2 full test)")
res_m2c = extract_coef_table(m2c, m2c_vars)

# M2d — Robustness: 50% threshold
m2d_vars = {k.replace("[H2: expect β₂ > β₁]","[robustness: 50% threshold]")
            .replace("[H2: comparison]",""):v for k,v in m2b_vars.items()}
m2d_formula = ("dependent_importers_50pct ~ "
               f"betweenness_L2_w + betweenness_L1_w + export_restriction_binary + "
               f"log_gdp {pop_term} + C(iso3) + C(year)")
m2d = run_ols_fe(df_exp, m2d_formula, "M2d 50% threshold (robustness)")
res_m2d = extract_coef_table(m2d, m2b_vars)

# ── Sub-period structural break check ─────────────────────────
print("\n─── Sub-period checks ───")
df_pre  = df_exp[df_exp["year"] <= 2018]
df_post = df_exp[df_exp["year"] >= 2019]

m2_pre  = run_ols_fe(df_pre,  m2b_formula, "M2 pre-2019")
m2_post = run_ols_fe(df_post, m2b_formula, "M2 post-2019")
res_m2_pre  = extract_coef_table(m2_pre,  m2b_vars)
res_m2_post = extract_coef_table(m2_post, m2b_vars)

# ── PRINT MAIN RESULTS ─────────────────────────────────────────
print("\n" + "═"*65)
print("MAIN RESULTS SUMMARY")
print("═"*65)

print("\nModel 2b — H2 test (Layer comparison):")
if m2b:
    bt_L2 = m2b.params.get("betweenness_L2_w", np.nan)
    bt_L1 = m2b.params.get("betweenness_L1_w", np.nan)
    p_L2  = m2b.pvalues.get("betweenness_L2_w", np.nan)
    p_L1  = m2b.pvalues.get("betweenness_L1_w", np.nan)
    print(f"  β(L2 midstream) = {bt_L2:.4f}  p={p_L2:.4f}")
    print(f"  β(L1 upstream)  = {bt_L1:.4f}  p={p_L1:.4f}")
    if bt_L2 > bt_L1:
        print(f"  → H2 SUPPORTED: midstream centrality coefficient ({bt_L2:.4f}) "
              f"> upstream ({bt_L1:.4f})")
    else:
        print(f"  → H2 NOT SUPPORTED in this spec. Check M2c.")

if m2c:
    print("\nModel 2c — Does centrality beat mining production?")
    bt_L2_c  = m2c.params.get("betweenness_L2_w", np.nan)
    bt_mine  = m2c.params.get("log_mining", np.nan)
    p_mine   = m2c.pvalues.get("log_mining", np.nan)
    p_L2_c   = m2c.pvalues.get("betweenness_L2_w", np.nan)
    print(f"  β(L2 betweenness) = {bt_L2_c:.4f}  p={p_L2_c:.4f}")
    print(f"  β(log mining)     = {bt_mine:.4f}  p={p_mine:.4f}")
    if p_L2_c < 0.10 and (not np.isfinite(p_mine) or p_mine > 0.10):
        print("  → H2 SUPPORTED: centrality significant, mining not significant after controlling for centrality")
    elif p_L2_c < 0.10 and p_mine < 0.10:
        print("  → Both significant — both mining and centrality matter (discuss in paper)")

# ══════════════════════════════════════════════════════════════
# SAVE TO EXCEL
# ══════════════════════════════════════════════════════════════
print("\nSaving regression tables to Excel...")
wb = Workbook()
wb.remove(wb.active)

def write_reg_sheet(wb, sheet_name, df_res, title, hypothesis, notes, tab_color="1F3864"):
    if df_res is None or len(df_res) == 0:
        return
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color

    DARK, GREY = "1F3864", "F2F2F2"

    # Title block
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font  = Font(bold=True, size=13, color=DARK)
    ws.cell(2, 1).value = f"Tests: {hypothesis}"
    ws.cell(2, 1).font  = Font(italic=True, size=10, color="7030A0")

    # Headers
    for c_idx, col in enumerate(df_res.columns, 1):
        cell = ws.cell(4, c_idx, value=col)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for r_idx, row in enumerate(df_res.itertuples(index=False), 5):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx,
                           value=None if isinstance(val, float) and np.isnan(val) else val)
            cell.alignment = Alignment(vertical="top")
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)
            if str(cell.value) in ("Country FE","Year FE","N","R² (within)","Adj. R²","Yes"):
                cell.font = Font(italic=True, size=9)

    # Notes
    note_row = len(df_res) + 6
    ws.cell(note_row, 1).value = f"Notes: {notes}"
    ws.cell(note_row, 1).font  = Font(italic=True, size=9, color="595959")
    ws.merge_cells(f"A{note_row}:{get_column_letter(len(df_res.columns))}{note_row}")
    ws.cell(note_row, 1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[note_row].height = 30

    ws.column_dimensions["A"].width = 50
    for c_ in range(2, len(df_res.columns)+1):
        ws.column_dimensions[get_column_letter(c_)].width = 14
    ws.freeze_panes = "B5"

FE_NOTE = "OLS with country (iso3) and year fixed effects. HC1 heteroskedasticity-robust standard errors. *** p<0.01 ** p<0.05 * p<0.10."

write_reg_sheet(wb, "M1a_Vulnerability_Main",
    res_m1a, "Model 1a — Supply Vulnerability (Main)",
    "H3: concentration → vulnerability; H4: MSP/alliance → less vulnerability",
    f"DV: China import share in Layer 2 (midstream REE compounds, HS 284690). {FE_NOTE}", "1F3864")

write_reg_sheet(wb, "M1b_Vuln_Plus_Centrality",
    res_m1b, "Model 1b — Vulnerability + Own Betweenness",
    "H3, H4 + network position effect",
    f"DV: China import share L2. Adds own betweenness centrality. {FE_NOTE}", "2E75B6")

write_reg_sheet(wb, "M1c_SVI_Network_Vars",
    res_m1c, "Model 1c — SVI with Network Variables Only",
    "H4: Coalition membership reduces composite vulnerability",
    f"DV: Supply Vulnerability Index (0-1). IVs: network position + policy vars (no SVI components). {FE_NOTE}", "4472C4")

write_reg_sheet(wb, "M2a_Bargaining_L2only",
    res_m2a, "Model 2a — Bargaining Power (L2 betweenness only, max N)",
    "H1: Betweenness L2 → more dependent importers",
    f"DV: Count of countries importing >25% of L2 REE from country i. {FE_NOTE}", "375623")

write_reg_sheet(wb, "M2b_Bargaining_L2vsL1",
    res_m2b, "Model 2b — Bargaining Power (L2 vs L1, H2 direct test)",
    "H2: β(L2 betweenness) > β(L1 betweenness) → midstream dominates mining",
    f"DV: Dependent importers (25% threshold). Direct test of H2. {FE_NOTE}", "70AD47")

write_reg_sheet(wb, "M2c_Plus_Mining",
    res_m2c, "Model 2c — Centrality vs Mining Production (H2 full test)",
    "H2: β(centrality) significant after controlling for log(mining production)",
    f"DV: Dependent importers (25% threshold). Smaller N due to USGS mining data coverage. {FE_NOTE}", "A9D18E")

write_reg_sheet(wb, "M2d_50pct_Robustness",
    res_m2d, "Model 2d — Robustness: 50% Dependence Threshold",
    "Robustness of H1 and H2 at stricter threshold",
    f"DV: Count of countries importing >50% of L2 REE from country i. {FE_NOTE}", "7030A0")

write_reg_sheet(wb, "M2_Pre2019",
    res_m2_pre, "Model 2 Sub-period: 2012–2018",
    "Structural stability check (pre-escalation period)",
    f"DV: Dependent importers (25%). Years 2012-2018 only. {FE_NOTE}", "C55A11")

write_reg_sheet(wb, "M2_Post2019",
    res_m2_post, "Model 2 Sub-period: 2019–2024",
    "Structural stability check (post-escalation: COVID + export controls)",
    f"DV: Dependent importers (25%). Years 2019-2024 only. {FE_NOTE}", "FF0000")

out_path = os.path.join(DATA_DIR, "05_regression_results.xlsx")
wb.save(out_path)
print(f"✓ Saved: {out_path}")
print("\n→ NEXT: run  Step6_Shock_Simulations.py")
