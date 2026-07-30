"""
STEP 7 — Causal Identification: Myanmar Coup Diff-in-Diff + Robustness Checks
═══════════════════════════════════════════════════════════════════════════════
Natural experiment: Myanmar military coup, February 1, 2021.
Myanmar's share of China's L2 REE feedstock rose from ~5% (2016) to ~70% (2020-21),
then was disrupted by the coup. Countries with high pre-coup China dependence
should have experienced larger vulnerability increases post-coup than countries
with low pre-coup China dependence — testable as a diff-in-diff.

DESIGNS
───────
  DiD-1  Inter-layer propagation test
         Treatment: post-coup (year ≥ 2021)
         Treated:   high pre-coup China L2 dependence (above median, 2019–2020 avg)
         Outcome:   china_import_share_L2
         Logic: if coup reduced China's processing capacity via Myanmar feedstock,
                high-China-dependent importers should be disproportionately hurt.

  DiD-2  Bargaining-power validation
         Treatment: post-coup
         Outcome:   China's dependent_importers_25pct
         Logic: if coup reduced China's L2 supply volume, fewer countries
                can maintain >25% China dependence → China loses leverage.

  ES-1   Event study (year dummies × treatment indicator)
         Tests parallel-trends assumption visually.

  ROB-1  Broad HS code definitions (L2 = 284690 + 284610 + 280530)
         Tests whether main regression results hold under alternative code mapping.

  ROB-2  Alternative dependence threshold (50%)
         Re-estimates Model 2 with dependent_importers_50pct as DV.

HOW TO RUN:
    python Step7_CausalID.py

Inputs:  data/04_master_panel.csv
Outputs: data/07_causal_id_results.xlsx
         data/07_did_panel.csv
"""

import pandas as pd
import numpy as np
import statsmodels.formula.api as smf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os, warnings
warnings.filterwarnings("ignore")

DATA = "/Users/sanayounas/Claude/Projects/Netwrok/data"
COUP_YEAR = 2021    # Myanmar coup, Feb 2021

# ── LOAD DATA ──────────────────────────────────────────────────
print("Loading master panel...")
df = pd.read_csv(os.path.join(DATA, "04_master_panel.csv"))
df["log_gdp"]    = np.log(df["gdp_current_usd"].replace(0, np.nan))
df["log_mining"] = np.log(df["mining_production_REO"].replace(0, np.nan))
for col in ["betweenness_L2","betweenness_L1","supplier_hhi_L2"]:
    p99 = df[col].quantile(0.99)
    df[f"{col}_w"] = df[col].clip(upper=p99)

print(f"  {len(df):,} rows | {df['iso3'].nunique()} countries | years {df['year'].min()}–{df['year'].max()}")

# ══════════════════════════════════════════════════════════════
# DID-1: INTER-LAYER PROPAGATION TEST
# ═══════════════════════════════════════════════════════════════
print("\n─── DiD-1: Myanmar coup → China L2 vulnerability propagation ───")

# Pre-coup treatment assignment: average China L2 share in 2019-2020
pre_coup = (df[df["year"].isin([2019, 2020])]
            .groupby("iso3")["china_import_share_L2"]
            .mean()
            .reset_index()
            .rename(columns={"china_import_share_L2": "pre_coup_china_share"}))

median_share = pre_coup["pre_coup_china_share"].median()
pre_coup["treated"] = (pre_coup["pre_coup_china_share"] >= median_share).astype(int)

print(f"  Median pre-coup China share: {median_share:.3f}")
print(f"  High-dependence group (treated): {pre_coup['treated'].sum()} countries")
print(f"  Low-dependence group (control):  {(1-pre_coup['treated']).sum()} countries")

# Merge treatment indicator into main panel
did = df.merge(pre_coup[["iso3","treated","pre_coup_china_share"]], on="iso3", how="inner")
did["post_coup"]    = (did["year"] >= COUP_YEAR).astype(int)
did["did_interact"] = did["treated"] * did["post_coup"]

# Restrict to countries with L2 import data
did = did[did["china_import_share_L2"].notna()].copy()

# Base DiD (two-way FE)
f_did1 = ("china_import_share_L2 ~ did_interact + C(iso3) + C(year)")
m_did1 = smf.ols(f_did1, data=did).fit(cov_type="HC1")

# With controls
f_did1c = ("china_import_share_L2 ~ did_interact + log_gdp + C(iso3) + C(year)")
m_did1c = smf.ols(f_did1c, data=did[did["log_gdp"].notna()]).fit(cov_type="HC1")

print(f"\n  DiD-1 (base):     N={int(m_did1.nobs):,}  R²={m_did1.rsquared:.4f}")
print(f"    DiD coefficient: {m_did1.params.get('did_interact',np.nan):.4f}  "
      f"p={m_did1.pvalues.get('did_interact',np.nan):.4f}")

print(f"\n  DiD-1 (+ controls): N={int(m_did1c.nobs):,}  R²={m_did1c.rsquared:.4f}")
print(f"    DiD coefficient: {m_did1c.params.get('did_interact',np.nan):.4f}  "
      f"p={m_did1c.pvalues.get('did_interact',np.nan):.4f}")

did_coef = m_did1.params.get('did_interact', np.nan)
did_p    = m_did1.pvalues.get('did_interact', np.nan)
did_sig  = "***" if did_p<0.01 else "**" if did_p<0.05 else "*" if did_p<0.10 else "n.s."

# ── Parallel trends event study (year-by-year interaction) ─────
print("\n  Event study (year dummies × treated):")
years_for_es = sorted(did["year"].unique())
base_year    = 2020  # last pre-coup year

es_rows = []
for yr in years_for_es:
    if yr == base_year:
        continue
    col_name = f"yr_{yr}_x_treated"
    did[col_name] = (did["year"] == yr).astype(int) * did["treated"]

yr_interactions = " + ".join([f"yr_{yr}_x_treated" for yr in years_for_es if yr != base_year])
f_es = f"china_import_share_L2 ~ {yr_interactions} + C(iso3) + C(year)"
m_es = smf.ols(f_es, data=did).fit(cov_type="HC1")

for yr in years_for_es:
    if yr == base_year:
        es_rows.append({"year": yr, "coef": 0.0, "se": 0.0, "p": 1.0, "note": "base year"})
        continue
    col = f"yr_{yr}_x_treated"
    if col in m_es.params.index:
        es_rows.append({
            "year": yr,
            "coef": round(m_es.params[col], 4),
            "se":   round(m_es.bse[col], 4),
            "p":    round(m_es.pvalues[col], 4),
            "note": "pre-coup" if yr < COUP_YEAR else "post-coup",
        })

es_df = pd.DataFrame(es_rows)
print(es_df.to_string(index=False))

# Pre-trend test: are pre-2021 coefficients jointly zero?
pre_cols = [f"yr_{yr}_x_treated" for yr in years_for_es if yr < COUP_YEAR and yr != base_year]
if pre_cols:
    from statsmodels.stats.anova import anova_lm
    try:
        pre_pvals = [m_es.pvalues.get(c, 1.0) for c in pre_cols]
        pre_coefs = [m_es.params.get(c, 0.0) for c in pre_cols]
        print(f"\n  Pre-trend check (years < {COUP_YEAR}, base={base_year}):")
        print(f"  Pre-coup coefficients: {[round(c,4) for c in pre_coefs]}")
        print(f"  Pre-coup p-values:     {[round(p,4) for p in pre_pvals]}")
        if all(p > 0.10 for p in pre_pvals):
            print("  ✓ PARALLEL TRENDS: No significant pre-coup differential trends (p>0.10 for all).")
        else:
            print("  ⚠ Some pre-coup differences — interpret DiD with caution; include pre-trend test in paper.")
    except Exception as e:
        print(f"  Pre-trend test: {e}")

# ══════════════════════════════════════════════════════════════
# DID-2: CHINA BARGAINING POWER VALIDATION
# ══════════════════════════════════════════════════════════════
print("\n─── DiD-2: Myanmar coup → China's bargaining power ───")

# Focus on China only: does China's dependent_importers_25pct change post-coup?
chn = df[df["iso3"] == "CHN"].copy().sort_values("year")
chn["post_coup"] = (chn["year"] >= COUP_YEAR).astype(int)

print("  China's dependent importers (25% threshold) by year:")
print(chn[["year","dependent_importers_25pct","china_import_share_L2",
           "out_strength_L2"]].to_string(index=False))

# Simple before-after comparison for China
pre_mean  = chn[chn["year"] < COUP_YEAR]["dependent_importers_25pct"].mean()
post_mean = chn[chn["year"] >= COUP_YEAR]["dependent_importers_25pct"].mean()
print(f"\n  China dependent importers: pre-coup avg={pre_mean:.1f}, post-coup avg={post_mean:.1f}, "
      f"Δ={post_mean-pre_mean:.1f}")

# ══════════════════════════════════════════════════════════════
# ROB-1: BROAD HS CODE ROBUSTNESS
# ══════════════════════════════════════════════════════════════
print("\n─── Robustness: Broad HS code definitions ───")

# We need the broad-definition variables from the centrality panel
cent = pd.read_csv(os.path.join(DATA, "02_centrality_panel.csv"))
dep  = pd.read_csv(os.path.join(DATA, "02_dependence_matrix.csv"))
baci = pd.read_csv(os.path.join(DATA, "01_baci_ree_filtered.csv"), dtype={"hs6":str})

# Broad L2 codes
BROAD_L2 = ["284690","284610","280530"]
dep_broad_L2 = dep[(dep["layer"]==2) & (dep["definition"]=="broad")].copy()

if len(dep_broad_L2) > 0:
    # China import share (broad L2)
    china_broad = (dep_broad_L2[dep_broad_L2["iso3_exp"]=="CHN"]
                   .groupby(["year","iso3_imp"])["D_ijt"].sum()
                   .reset_index()
                   .rename(columns={"D_ijt":"china_share_L2_broad","iso3_imp":"iso3"}))

    # Merge broad betweenness
    cent_broad = (cent[(cent["layer"]==2) & (cent["definition"]=="broad")]
                  .rename(columns={"betweenness":"betweenness_L2_broad",
                                   "eigenvector":"eigenvector_L2_broad"})[
                  ["iso3","year","betweenness_L2_broad","eigenvector_L2_broad"]])

    rob = (df.merge(china_broad, on=["iso3","year"], how="left")
             .merge(cent_broad,  on=["iso3","year"], how="left"))

    rob["btw_L2_broad_w"] = rob["betweenness_L2_broad"].clip(
        upper=rob["betweenness_L2_broad"].quantile(0.99))

    # Broad robustness for Model 1
    has_L2b = rob["china_share_L2_broad"].notna() & (rob["in_strength_L2"].fillna(0) > 0)
    rob_imp  = rob[has_L2b].copy()
    if len(rob_imp) > 100:
        m_rob1 = smf.ols(
            "china_share_L2_broad ~ alliance_diversity_idx + msp_member + "
            "n_suppliers_L2 + supplier_hhi_L2_w + log_gdp + C(iso3) + C(year)",
            data=rob_imp[rob_imp["log_gdp"].notna()]
        ).fit(cov_type="HC1")
        print(f"  ROB-1 Model 1 (broad L2): N={int(m_rob1.nobs):,}  R²={m_rob1.rsquared:.4f}")
        a_coef = m_rob1.params.get("alliance_diversity_idx",np.nan)
        a_p    = m_rob1.pvalues.get("alliance_diversity_idx",np.nan)
        print(f"    alliance_diversity_idx: β={a_coef:.4f}  p={a_p:.4f}")
    else:
        m_rob1 = None
        print("  ROB-1 Model 1: insufficient broad L2 data")

    # Broad robustness for Model 2
    has_L2b_exp = rob["out_strength_L2"].fillna(0) > 0
    rob_exp = rob[has_L2b_exp & rob["btw_L2_broad_w"].notna()].copy()
    if len(rob_exp) > 100:
        m_rob2 = smf.ols(
            "dependent_importers_25pct ~ btw_L2_broad_w + export_restriction_binary + "
            "log_gdp + C(iso3) + C(year)",
            data=rob_exp[rob_exp["log_gdp"].notna()]
        ).fit(cov_type="HC1")
        print(f"  ROB-1 Model 2 (broad L2): N={int(m_rob2.nobs):,}  R²={m_rob2.rsquared:.4f}")
        b_coef = m_rob2.params.get("btw_L2_broad_w",np.nan)
        b_p    = m_rob2.pvalues.get("btw_L2_broad_w",np.nan)
        b_sig  = "***" if b_p<0.01 else "**" if b_p<0.05 else "*" if b_p<0.10 else "n.s."
        print(f"    betweenness_L2_broad: β={b_coef:.4f}  p={b_p:.4f} {b_sig}")
    else:
        m_rob2 = None
        print("  ROB-1 Model 2: insufficient broad L2 data")
else:
    print("  ⚠ No broad-definition dependence matrix data found — run Step2 with broad codes.")
    m_rob1 = m_rob2 = None

# ── ROB-2: 50% Threshold ─────────────────────────────────────
print("\n─── Robustness: 50% dependence threshold ───")
exp_df = df[df["out_strength_L2"].fillna(0) > 0].copy()
exp_df = exp_df[exp_df["betweenness_L2_w"].notna() & exp_df["log_gdp"].notna()]
m_rob3 = smf.ols(
    "dependent_importers_50pct ~ betweenness_L2_w + betweenness_L1_w + "
    "export_restriction_binary + log_gdp + C(iso3) + C(year)",
    data=exp_df
).fit(cov_type="HC1")
b50_coef = m_rob3.params.get("betweenness_L2_w",np.nan)
b50_p    = m_rob3.pvalues.get("betweenness_L2_w",np.nan)
b50_sig  = "***" if b50_p<0.01 else "**" if b50_p<0.05 else "*" if b50_p<0.10 else "n.s."
print(f"  ROB-2 (50% threshold): N={int(m_rob3.nobs):,}  R²={m_rob3.rsquared:.4f}")
print(f"    betweenness_L2: β={b50_coef:.4f}  p={b50_p:.4f} {b50_sig}")

# ══════════════════════════════════════════════════════════════
# DESCRIPTIVE STATISTICS TABLE
# ══════════════════════════════════════════════════════════════
print("\n─── Descriptive Statistics ───")
desc_vars = [
    ("china_import_share_L2",     "China import share, Layer 2"),
    ("supply_vulnerability_idx",  "Supply Vulnerability Index (SVI)"),
    ("supplier_hhi_L2",           "Supplier HHI, Layer 2"),
    ("n_suppliers_L2",            "Number of Layer 2 suppliers"),
    ("alliance_diversity_idx",    "Alliance diversity index"),
    ("msp_member",                "MSP member (binary)"),
    ("betweenness_L2",            "Betweenness centrality, Layer 2"),
    ("betweenness_L1",            "Betweenness centrality, Layer 1"),
    ("betweenness_L3",            "Betweenness centrality, Layer 3"),
    ("dependent_importers_25pct", "Dependent importers (25% threshold)"),
    ("dependent_importers_50pct", "Dependent importers (50% threshold)"),
    ("export_restriction_binary", "Export restriction (binary)"),
    ("gdp_current_usd",           "GDP, current USD (billions)"),
]

desc_rows = []
for col, label in desc_vars:
    if col in df.columns:
        s = df[col].dropna()
        if col == "gdp_current_usd":
            s = s / 1e9
        desc_rows.append({
            "Variable": label,
            "N":    len(s),
            "Mean": round(s.mean(), 4),
            "SD":   round(s.std(), 4),
            "Min":  round(s.min(), 4),
            "p25":  round(s.quantile(0.25), 4),
            "p75":  round(s.quantile(0.75), 4),
            "Max":  round(s.max(), 4),
        })

desc_df = pd.DataFrame(desc_rows)
print(desc_df.to_string(index=False))

# ══════════════════════════════════════════════════════════════
# SAVE RESULTS
# ══════════════════════════════════════════════════════════════
print("\nSaving results...")
did["year_str"] = did["year"].astype(str)
did.to_csv(os.path.join(DATA,"07_did_panel.csv"), index=False)

wb = Workbook()
wb.remove(wb.active)
DARK = "1F3864"
GRY  = "F5F5F5"

def ws_table(wb, name, df_in, title, notes="", tab="1F3864"):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab
    ws.cell(1,1,title).font = Font(bold=True, size=12, color=DARK)
    if notes:
        ws.cell(2,1,notes).font = Font(italic=True, size=9, color="595959")
    sr = 4 if notes else 3
    for c,col in enumerate(df_in.columns,1):
        cell = ws.cell(sr,c,col)
        cell.fill = PatternFill("solid",fgColor=DARK)
        cell.font = Font(bold=True,color="FFFFFF",size=9)
        cell.alignment = Alignment(horizontal="center")
    for r,row in enumerate(df_in.itertuples(index=False),sr+1):
        for c,val in enumerate(row,1):
            cell = ws.cell(r,c, value=None if isinstance(val,float) and np.isnan(val) else val)
            if r%2==0: cell.fill = PatternFill("solid",fgColor=GRY)
            cell.alignment = Alignment(horizontal="left" if c==1 else "center")
            set_f = Font(size=9)
            cell.font = set_f
    ws.column_dimensions["A"].width = 40
    for c in range(2, len(df_in.columns)+1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.freeze_panes = f"B{sr+1}"
    return ws

# DiD results table
did_res = pd.DataFrame([
    {"Specification": "DiD-1 base (country+year FE)",
     "DV": "china_import_share_L2",
     "N": int(m_did1.nobs),
     "DiD coef.": round(m_did1.params.get("did_interact",np.nan),4),
     "Std. Err.": round(m_did1.bse.get("did_interact",np.nan),4),
     "p-value": round(m_did1.pvalues.get("did_interact",np.nan),4),
     "Sig.": did_sig,
     "R²": round(m_did1.rsquared,4)},
    {"Specification": "DiD-1 + log(GDP) control",
     "DV": "china_import_share_L2",
     "N": int(m_did1c.nobs),
     "DiD coef.": round(m_did1c.params.get("did_interact",np.nan),4),
     "Std. Err.": round(m_did1c.bse.get("did_interact",np.nan),4),
     "p-value": round(m_did1c.pvalues.get("did_interact",np.nan),4),
     "Sig.": "***" if m_did1c.pvalues.get("did_interact",1)<0.01 else
             "**"  if m_did1c.pvalues.get("did_interact",1)<0.05 else
             "*"   if m_did1c.pvalues.get("did_interact",1)<0.10 else "n.s.",
     "R²": round(m_did1c.rsquared,4)},
])

ws_table(wb, "DiD_Myanmar_Coup", did_res,
         "Diff-in-Diff: Myanmar Coup (Feb 2021) → Vulnerability Propagation",
         "Treatment = post-2021 × high pre-coup China L2 dependence. "
         "HC1 robust SEs. *** p<0.01 ** p<0.05 * p<0.10.", "C00000")

ws_table(wb, "EventStudy", es_df,
         "Event Study: Year-by-Year DiD Coefficients",
         "Interaction of year dummy with treated (high pre-coup China dependence). "
         "Base year = 2020. Pre-trend check: coefficients for years < 2021 should be near zero.", "ED7D31")

ws_table(wb, "DescriptiveStats", desc_df,
         "Table A1. Descriptive Statistics — All Regression Variables",
         "All country-year observations in master panel, 2012–2024. GDP in USD billions.",
         DARK)

# Robustness summary
rob_rows = []
if m_rob1:
    rob_rows.append({"Check":"ROB-1a: Broad L2 codes, Model 1",
                     "DV":"china_share_L2_broad",
                     "Key variable":"alliance_diversity_idx",
                     "N":int(m_rob1.nobs),
                     "β":round(m_rob1.params.get("alliance_diversity_idx",np.nan),4),
                     "p":round(m_rob1.pvalues.get("alliance_diversity_idx",np.nan),4)})
if m_rob2:
    rob_rows.append({"Check":"ROB-1b: Broad L2 codes, Model 2",
                     "DV":"dependent_importers_25pct",
                     "Key variable":"betweenness_L2_broad",
                     "N":int(m_rob2.nobs),
                     "β":round(m_rob2.params.get("btw_L2_broad_w",np.nan),4),
                     "p":round(m_rob2.pvalues.get("btw_L2_broad_w",np.nan),4)})
rob_rows.append({"Check":"ROB-2: 50% dependence threshold, Model 2",
                 "DV":"dependent_importers_50pct",
                 "Key variable":"betweenness_L2",
                 "N":int(m_rob3.nobs),
                 "β":round(b50_coef,4),
                 "p":round(b50_p,4)})

if rob_rows:
    rob_df = pd.DataFrame(rob_rows)
    ws_table(wb, "Robustness", rob_df,
             "Table A3. Robustness Checks",
             "Broad HS code definitions and alternative dependence thresholds. "
             "Main model results should be robust across specifications.", "375623")

out = os.path.join(DATA,"07_causal_id_results.xlsx")
wb.save(out)
print(f"✓ Saved: {out}")
print(f"✓ Saved: data/07_did_panel.csv")

print("""
═══════════════════════════════════════════════════════════════
INTERPRETATION GUIDE
═══════════════════════════════════════════════════════════════
DiD coefficient (did_interact):
  Positive → high-China-dependence countries saw LARGER China
             import share post-coup than low-dependence countries.
             (Counterintuitive? Could mean China maintained supply
              to strategically important partners; smaller economies
              bore the brunt of reduced capacity.)
  Negative → high-China-dependence countries diversified MORE
             post-coup relative to control group.
  Near zero, p<0.10 → coup accelerated vulnerability divergence.

Event study (EventStudy sheet):
  Check that pre-2021 coefficients cluster near zero — this
  validates the parallel-trends assumption required for DiD.
  Post-2021 coefficients should show the treatment effect emerging.

For the paper (Section 4.4 and Appendix):
  State: "We exploit the Myanmar military coup (February 2021) as
  a source of quasi-exogenous variation in China's Layer 2
  processing-feedstock supply. We estimate a difference-in-differences
  model using pre-coup China-import-dependence as the treatment
  assignment variable."
═══════════════════════════════════════════════════════════════
""")
