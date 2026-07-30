"""
STEP 6 — Supply-Chain Shock Simulations
════════════════════════════════════════
Simulates four counterfactual scenarios on the 2023 REE trade network
and measures how vulnerability and bargaining power change.

SCENARIOS
─────────
  S1  China L2 export ban     — remove all CHN edges in Layer 2 (midstream)
  S2  Myanmar L1 disruption   — remove all MMR edges in Layer 1 (upstream ores)
  S3  Dual shock              — S1 + S2 simultaneously
  S4  MSP alliance scale-up   — double L2 export flows from all MSP members
                                 (positive diversification scenario)

METRICS TRACKED (per country, before vs after)
  • china_import_share_L2   — direct China dependence
  • supplier_hhi_L2         — concentration of remaining suppliers
  • n_suppliers_L2          — number of active suppliers
  • supply_vulnerability_idx (SVI) — composite 0–1 index
  • dependent_importers_25pct — exporter-side bargaining power

HOW TO RUN:
    python Step6_Shock_Simulations.py

Inputs:  data/01_baci_ree_filtered.csv
         data/04_master_panel.csv
Outputs: data/06_shock_results.csv
         data/06_shock_summary.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os, warnings
warnings.filterwarnings("ignore")

DATA_DIR    = "/Users/sanayounas/Claude/Projects/Netwrok/data"
PROJECT_DIR = "/Users/sanayounas/Claude/Projects/Netwrok"

# ── CONFIGURATION ──────────────────────────────────────────────
SHOCK_YEAR    = 2023      # most recent complete year
NARROW_L1     = ["261790"]
NARROW_L2     = ["284690"]
NARROW_L3     = ["850511"]
MSP_MEMBERS   = {"USA","CAN","AUS","GBR","FRA","DEU","JPN","KOR","IND","ITA"}
HIGH_VULN_THR = 0.70      # SVI threshold for "high vulnerability" flag
DEP_THR       = 0.25      # dependence threshold for bargaining power count

# ── LOAD DATA ──────────────────────────────────────────────────
print("Loading data...")
baci = pd.read_csv(os.path.join(DATA_DIR, "01_baci_ree_filtered.csv"),
                   dtype={"hs6": str})
master = pd.read_csv(os.path.join(DATA_DIR, "04_master_panel.csv"))

# Filter to narrow codes, shock year
baci_yr = baci[baci["year"] == SHOCK_YEAR].copy()
baci_L1 = baci_yr[baci_yr["hs6"].isin(NARROW_L1)].copy()
baci_L2 = baci_yr[baci_yr["hs6"].isin(NARROW_L2)].copy()
baci_L3 = baci_yr[baci_yr["hs6"].isin(NARROW_L3)].copy()

print(f"  Shock year {SHOCK_YEAR}: "
      f"L1={len(baci_L1):,} flows, L2={len(baci_L2):,} flows, L3={len(baci_L3):,} flows")

# ── HELPER FUNCTIONS ───────────────────────────────────────────

def compute_vuln_metrics(edges_L2):
    """
    Given a Layer 2 edge dataframe (iso3_exp, iso3_imp, value_usd),
    compute vulnerability metrics for each importer.
    """
    if len(edges_L2) == 0:
        return pd.DataFrame(columns=["iso3","china_import_share_L2",
                                      "supplier_hhi_L2","n_suppliers_L2",
                                      "supply_vulnerability_idx"])

    agg = edges_L2.groupby(["iso3_exp","iso3_imp"])["value_usd"].sum().reset_index()
    agg = agg[agg["value_usd"] > 0].copy()

    total_imp = agg.groupby("iso3_imp")["value_usd"].transform("sum")
    agg["D_ijt"] = agg["value_usd"] / total_imp

    # China share
    china = (agg[agg["iso3_exp"] == "CHN"]
             .groupby("iso3_imp")["D_ijt"].sum()
             .reset_index()
             .rename(columns={"D_ijt": "china_import_share_L2", "iso3_imp": "iso3"}))

    # HHI
    hhi = (agg.groupby("iso3_imp")
           .apply(lambda x: ((x["D_ijt"] * 100) ** 2).sum(), include_groups=False)
           .reset_index()
           .rename(columns={0: "supplier_hhi_L2", "iso3_imp": "iso3"}))

    # N suppliers
    n_sup = (agg[agg["value_usd"] > 0]
             .groupby("iso3_imp")["iso3_exp"].nunique()
             .reset_index()
             .rename(columns={"iso3_exp": "n_suppliers_L2", "iso3_imp": "iso3"}))

    # Base: all importers
    base = (agg.groupby("iso3_imp")["value_usd"].sum().reset_index()
            .rename(columns={"iso3_imp":"iso3","value_usd":"total_L2_imports"}))

    v = (base
         .merge(china, on="iso3", how="left")
         .merge(hhi,   on="iso3", how="left")
         .merge(n_sup, on="iso3", how="left"))

    v["china_import_share_L2"] = v["china_import_share_L2"].fillna(0)
    v["supplier_hhi_L2"]        = v["supplier_hhi_L2"].fillna(0)
    v["n_suppliers_L2"]         = v["n_suppliers_L2"].fillna(0)

    # SVI
    max_sup = v["n_suppliers_L2"].replace(0, np.nan).max()
    v["n_sup_norm"] = 1 - (v["n_suppliers_L2"] / max_sup)
    v["supply_vulnerability_idx"] = (
        0.40 * v["china_import_share_L2"].clip(0, 1) +
        0.30 * (v["supplier_hhi_L2"] / 10000).clip(0, 1) +
        0.30 * v["n_sup_norm"].fillna(1)
    ).round(4)

    return v[["iso3","china_import_share_L2","supplier_hhi_L2",
              "n_suppliers_L2","supply_vulnerability_idx"]]


def compute_barg_metrics(edges_L2):
    """Count dependent importers (>25%) per exporter."""
    if len(edges_L2) == 0:
        return pd.DataFrame(columns=["iso3","dependent_importers_25pct"])

    agg = edges_L2.groupby(["iso3_exp","iso3_imp"])["value_usd"].sum().reset_index()
    agg = agg[agg["value_usd"] > 0].copy()
    total_imp = agg.groupby("iso3_imp")["value_usd"].transform("sum")
    agg["D_ijt"] = agg["value_usd"] / total_imp

    dep = (agg[agg["D_ijt"] >= DEP_THR]
           .groupby("iso3_exp")["iso3_imp"].nunique()
           .reset_index()
           .rename(columns={"iso3_imp": "dependent_importers_25pct", "iso3_exp": "iso3"}))
    return dep


def apply_shock_L2(edges, remove_exporters=None, scale_exporters=None, scale_factor=2.0):
    """
    Modify Layer 2 edges:
      remove_exporters: list of iso3 to remove entirely
      scale_exporters:  list of iso3 whose flows get multiplied by scale_factor
    """
    df = edges.copy()
    if remove_exporters:
        df = df[~df["iso3_exp"].isin(remove_exporters)]
    if scale_exporters:
        mask = df["iso3_exp"].isin(scale_exporters)
        df.loc[mask, "value_usd"] *= scale_factor
    return df


def apply_shock_L1(edges, remove_exporters=None):
    df = edges.copy()
    if remove_exporters:
        df = df[~df["iso3_exp"].isin(remove_exporters)]
    return df


def propagate_L2_feedstock_shock(edges_L2_baseline, disrupted_supplier, processor="CHN"):
    """
    Propagate a disruption of one country's L2 exports TO the processor into
    the processor's L2 exports TO the rest of the world.

    Why L2-to-L2 (not L1-to-L2):
      Myanmar exports HS 284690 (rare earth compounds) directly to China — these
      are classified as Layer 2 in BACI trade data, not Layer 1 ores. Myanmar's
      role is as a primary L2 feedstock supplier to China's processing/re-export
      network. In 2023 Myanmar supplies ~65.9% of China's L2 import feedstock.

    Mechanism:
      1. Compute disrupted_supplier's share of processor's total L2 imports.
      2. Assume processor's L2 export capacity is proportional to feedstock available.
      3. Scale processor's L2 exports to all destinations by (1 − feedstock_loss_share).

    Note: This is an upper-bound impact assumption — it assumes no domestic buffer
    stock or alternative sourcing. State this limitation in the paper.

    Arguments:
      edges_L2_baseline   : full L2 trade network
      disrupted_supplier  : iso3 removed from L2 supply to processor (e.g. "MMR")
      processor           : iso3 whose exports are scaled down (default "CHN")

    Returns: modified L2 edges with processor's exports scaled down.
    """
    # Processor's total L2 imports from all sources in baseline
    proc_imports_bl = edges_L2_baseline[edges_L2_baseline["iso3_imp"] == processor]
    total_L2_in_bl  = proc_imports_bl["value_usd"].sum()

    # How much did disrupted_supplier provide?
    supplier_to_proc = proc_imports_bl[proc_imports_bl["iso3_exp"] == disrupted_supplier]["value_usd"].sum()

    if total_L2_in_bl == 0:
        print(f"  ⚠ {processor} has no L2 imports in baseline — no propagation applied.")
        return edges_L2_baseline.copy()

    feedstock_loss_share = supplier_to_proc / total_L2_in_bl
    scale_factor         = max(0.0, 1.0 - feedstock_loss_share)

    print(f"  L2 feedstock propagation: {disrupted_supplier} → {processor} → world")
    print(f"    {processor} total L2 imports (baseline): ${total_L2_in_bl/1e6:,.1f}M")
    print(f"    {disrupted_supplier} share of {processor} L2 imports: {feedstock_loss_share*100:.1f}%")
    print(f"    → {processor} L2 export scale factor: {scale_factor:.3f}")
    print(f"      (Upper-bound: assumes no domestic buffer or alternative sourcing)")

    # Remove the disrupted_supplier → processor flow from the network
    edges_L2 = edges_L2_baseline.copy()
    edges_L2 = edges_L2[~((edges_L2["iso3_exp"] == disrupted_supplier) &
                           (edges_L2["iso3_imp"] == processor))]

    # Scale processor's exports to the rest of the world
    mask = edges_L2["iso3_exp"] == processor
    edges_L2.loc[mask, "value_usd"] *= scale_factor

    return edges_L2


# ── BASELINE (2023, no shock) ──────────────────────────────────
print("\nComputing baseline metrics (2023)...")
bl_vuln = compute_vuln_metrics(baci_L2)
bl_barg = compute_barg_metrics(baci_L2)
baseline = bl_vuln.merge(bl_barg, on="iso3", how="outer")
baseline["dependent_importers_25pct"] = baseline["dependent_importers_25pct"].fillna(0).astype(int)

print(f"  Baseline countries: {len(baseline)}")
print(f"  China L2 share (global avg): {bl_vuln['china_import_share_L2'].mean():.3f}")
print(f"  Countries with SVI > {HIGH_VULN_THR}: {(bl_vuln['supply_vulnerability_idx'] > HIGH_VULN_THR).sum()}")


# ── PRE-COMPUTE PROPAGATED L2 EDGES FOR MYANMAR SCENARIOS ─────
print("\nPre-computing Myanmar feedstock propagation (S2 and S3)...")

# Key finding from data: Myanmar exports HS 284690 (L2 compounds) directly to China,
# not L1 ores. Myanmar supplies ~65.9% of China's L2 import feedstock in 2023.
# Shock: remove Myanmar as China's L2 feedstock supplier → scale China's L2 exports down.

baci_L2_s2 = propagate_L2_feedstock_shock(
    edges_L2_baseline=baci_L2,
    disrupted_supplier="MMR",
    processor="CHN"
)

# S3: Myanmar disruption AND China L2 ban
# Apply Myanmar feedstock shock first (scales China exports), then remove China entirely
# Net effect: China is already removed by the ban, so S3 ≈ S1 + Myanmar loses L2 route to China
# The more meaningful dual shock: Myanmar removed from L2 to China, China exports banned
baci_L2_s3 = apply_shock_L2(baci_L2_s2, remove_exporters=["CHN"])

# ── RUN ALL FOUR SCENARIOS ─────────────────────────────────────
scenarios = {
    "S1_China_L2_ban":      ("China L2 export ban",
                              apply_shock_L2(baci_L2, remove_exporters=["CHN"])),
    "S2_Myanmar_L1_shock":  ("Myanmar L1 disruption → China L2 capacity reduction",
                              baci_L2_s2),
    "S3_Dual_shock":        ("Dual shock (China L2 ban + Myanmar L1 disruption)",
                              baci_L2_s3),
    "S4_MSP_scale_up":      ("MSP alliance scale-up (×2 exports)",
                              apply_shock_L2(baci_L2, scale_exporters=list(MSP_MEMBERS), scale_factor=2.0)),
}

all_results = []

for sid, (label, edges_L2_shock) in scenarios.items():
    print(f"\n{sid}: {label}")

    s_vuln = compute_vuln_metrics(edges_L2_shock)
    s_barg = compute_barg_metrics(edges_L2_shock)
    s_df   = s_vuln.merge(s_barg, on="iso3", how="outer")
    s_df["dependent_importers_25pct"] = s_df["dependent_importers_25pct"].fillna(0).astype(int)

    # Merge with baseline
    merged = baseline.merge(s_df, on="iso3", how="outer", suffixes=("_bl","_shock"))

    # Compute deltas
    for col in ["china_import_share_L2","supplier_hhi_L2","n_suppliers_L2",
                "supply_vulnerability_idx","dependent_importers_25pct"]:
        bl_c = f"{col}_bl"
        sh_c = f"{col}_shock"
        if bl_c in merged.columns and sh_c in merged.columns:
            merged[f"delta_{col}"] = merged[sh_c] - merged[bl_c]

    # Flag high vulnerability
    merged["high_vuln_baseline"] = (merged["supply_vulnerability_idx_bl"] > HIGH_VULN_THR).astype(int)
    merged["high_vuln_shock"]    = (merged["supply_vulnerability_idx_shock"] > HIGH_VULN_THR).astype(int)
    merged["newly_vulnerable"]   = ((merged["high_vuln_shock"] == 1) & (merged["high_vuln_baseline"] == 0)).astype(int)

    merged["scenario_id"]    = sid
    merged["scenario_label"] = label
    all_results.append(merged)

    # Print summary
    n_new = merged["newly_vulnerable"].sum()
    avg_delta_svi = merged["delta_supply_vulnerability_idx"].mean()
    print(f"  Newly high-vulnerability countries: {n_new}")
    print(f"  Avg ΔSVI across all importers: {avg_delta_svi:+.4f}")

    if sid == "S1_China_L2_ban":
        worst = merged.nlargest(5, "delta_supply_vulnerability_idx")[["iso3","supply_vulnerability_idx_bl","supply_vulnerability_idx_shock","delta_supply_vulnerability_idx"]]
        print(f"  Top 5 hardest hit:")
        print(worst.to_string(index=False))

    if sid == "S4_MSP_scale_up":
        best = merged.nsmallest(5, "delta_supply_vulnerability_idx")[["iso3","supply_vulnerability_idx_bl","supply_vulnerability_idx_shock","delta_supply_vulnerability_idx"]]
        print(f"  Top 5 most improved:")
        print(best.to_string(index=False))


# ── SAVE RESULTS ───────────────────────────────────────────────
print("\nSaving shock results...")
results_df = pd.concat(all_results, ignore_index=True)
results_df.to_csv(os.path.join(DATA_DIR, "06_shock_results.csv"), index=False)
print(f"  ✓ Saved: data/06_shock_results.csv ({len(results_df):,} rows)")


# ── BUILD COMPARISON TABLE ─────────────────────────────────────
print("\nBuilding comparison table...")
compare_rows = []
for sid, (label, _) in scenarios.items():
    sc = results_df[results_df["scenario_id"] == sid].copy()
    compare_rows.append({
        "Scenario": label,
        "Avg ΔSVI": round(sc["delta_supply_vulnerability_idx"].mean(), 4),
        "Max ΔSVI": round(sc["delta_supply_vulnerability_idx"].max(), 4),
        "Newly high-vuln countries": int(sc["newly_vulnerable"].sum()),
        "Countries losing China supply": int((sc["delta_china_import_share_L2"] < -0.10).sum()),
        "Avg Δ n_suppliers": round(sc["delta_n_suppliers_L2"].mean(), 2),
        "Avg Δ HHI": round(sc["delta_supplier_hhi_L2"].mean(), 1),
    })
compare_df = pd.DataFrame(compare_rows)

print("\nSCENARIO COMPARISON TABLE:")
print(compare_df.to_string(index=False))


# ── WHO GAINS BARGAINING POWER AFTER S1? ──────────────────────
print("\nBargaining power shift after China ban (S1):")
s1 = results_df[results_df["scenario_id"] == "S1_China_L2_ban"].copy()
s1_barg = s1[["iso3","dependent_importers_25pct_bl","dependent_importers_25pct_shock",
               "delta_dependent_importers_25pct"]].dropna()
s1_barg = s1_barg.sort_values("delta_dependent_importers_25pct", ascending=False)
print("  Countries gaining most dependent importers post-China ban:")
print(s1_barg[s1_barg["delta_dependent_importers_25pct"] > 0].head(10).to_string(index=False))


# ── WRITE EXCEL WORKBOOK ───────────────────────────────────────
print("\nWriting Excel output...")
wb = Workbook()
wb.remove(wb.active)

DARK = "1F3864"
GREY = "F2F2F2"
RED  = "C00000"
GRN  = "375623"
AMB  = "C55A11"

def write_sheet(wb, name, df, title, subtitle="", tab_color=DARK):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    ws.cell(1, 1, title).font = Font(bold=True, size=12, color=DARK)
    if subtitle:
        ws.cell(2, 1, subtitle).font = Font(italic=True, size=9, color="595959")
    start_row = 3 if not subtitle else 4

    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(start_row, c_idx, col)
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx,
                           value=None if isinstance(val, float) and np.isnan(val) else val)
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)
            cell.alignment = Alignment(vertical="top")

    ws.column_dimensions["A"].width = 20
    for c_ in range(2, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(c_)].width = 16
    ws.freeze_panes = f"B{start_row + 1}"


# Sheet 1 — Summary comparison
write_sheet(wb, "Summary", compare_df,
            "Scenario Comparison — Key Impact Metrics",
            f"Shock year: {SHOCK_YEAR}. Layer 2 narrow definition (HS 284690).",
            DARK)

# Sheets 2–5 — Per-scenario detailed results
scenario_colors = {"S1_China_L2_ban": RED, "S2_Myanmar_L1_shock": AMB,
                   "S3_Dual_shock": "7030A0", "S4_MSP_scale_up": GRN}

for sid, (label, _) in scenarios.items():
    sc = results_df[results_df["scenario_id"] == sid].copy()
    cols_to_show = [
        "iso3",
        "china_import_share_L2_bl",  "china_import_share_L2_shock",  "delta_china_import_share_L2",
        "supplier_hhi_L2_bl",         "supplier_hhi_L2_shock",         "delta_supplier_hhi_L2",
        "n_suppliers_L2_bl",          "n_suppliers_L2_shock",          "delta_n_suppliers_L2",
        "supply_vulnerability_idx_bl","supply_vulnerability_idx_shock","delta_supply_vulnerability_idx",
        "dependent_importers_25pct_bl","dependent_importers_25pct_shock","delta_dependent_importers_25pct",
        "newly_vulnerable",
    ]
    cols_present = [c for c in cols_to_show if c in sc.columns]
    sc_out = (sc[cols_present]
              .sort_values("delta_supply_vulnerability_idx", ascending=False)
              .reset_index(drop=True))
    for col in sc_out.select_dtypes(include=float).columns:
        sc_out[col] = sc_out[col].round(4)

    write_sheet(wb, sid[:25], sc_out, f"{sid}: {label}",
                f"Country-level changes vs 2023 baseline. Sorted by ΔSVI (worst hit first).",
                scenario_colors.get(sid, DARK))

# Sheet 6 — Bargaining power shift under S1
s1_bp = (results_df[results_df["scenario_id"] == "S1_China_L2_ban"]
         [["iso3","dependent_importers_25pct_bl","dependent_importers_25pct_shock",
           "delta_dependent_importers_25pct"]]
         .dropna(subset=["delta_dependent_importers_25pct"])
         .sort_values("delta_dependent_importers_25pct", ascending=False)
         .reset_index(drop=True))
s1_bp["delta_dependent_importers_25pct"] = s1_bp["delta_dependent_importers_25pct"].round(0).astype(int)
write_sheet(wb, "S1_BargainingPowerShift", s1_bp,
            "S1: Who gains bargaining power after China L2 ban?",
            "Countries sorted by change in dependent importers (25% threshold).",
            RED)

out_path = os.path.join(DATA_DIR, "06_shock_summary.xlsx")
wb.save(out_path)
print(f"✓ Saved: {out_path}")

print("""
═══════════════════════════════════════════════════════════
DONE. Key outputs:
  data/06_shock_results.csv     — full country × scenario panel
  data/06_shock_summary.xlsx    — formatted tables (6 sheets)

Interpretation guide:
  S1 (China ban): identifies which countries are most exposed
                  and which alternative suppliers gain leverage.
  S2 (Myanmar):   upstream disruption impact on China's own
                  processing inputs — indirect vulnerability.
  S3 (Dual):      worst-case combined shock — use as ceiling.
  S4 (MSP ×2):    best-case diversification — use as policy target.

→ NEXT: draft paper sections using these results.
═══════════════════════════════════════════════════════════
""")
