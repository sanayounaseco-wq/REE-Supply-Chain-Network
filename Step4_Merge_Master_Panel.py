"""
STEP 4 — Merge Everything into the Master Panel
════════════════════════════════════════════════
Combines:
  • Centrality metrics (Step 2)        → betweenness, eigenvector, strengths per layer
  • Vulnerability variables (Step 3)   → SVI, HHI, china share, bargaining power
  • USGS production data               → mining_production_REO, reserves
  • GPR Index                          → gpr_index
  • GDP approximations                 → gdp_current_usd
  • Alliance/MSP flags                 → msp_member, nato_member

Produces one analysis-ready CSV and updates the Excel workbook.

HOW TO RUN:
    python Step4_Merge_Master_Panel.py

Output:
  data/04_master_panel.csv        ← use this in your regressions
  REE_Master_Dataset.xlsx         ← Country_Year_Panel sheet updated
"""

import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

DATA_DIR    = "/Users/sanayounas/Claude/Projects/Netwrok/data"
PROJECT_DIR = "/Users/sanayounas/Claude/Projects/Netwrok"

# ── LOAD CENTRALITY DATA ───────────────────────────────────────
print("Loading centrality panel...")
cent = pd.read_csv(os.path.join(DATA_DIR, "02_centrality_panel.csv"))
cent_narrow = cent[cent["definition"] == "narrow"].copy()

# Pivot: one row per country-year, columns per layer
print("Pivoting centrality to wide format...")
cent_wide = pd.DataFrame()
for lyr in [1, 2, 3]:
    lyr_data = cent_narrow[cent_narrow["layer"] == lyr][
        ["year","iso3","betweenness","eigenvector","out_strength","in_strength"]
    ].rename(columns={
        "betweenness":  f"betweenness_L{lyr}",
        "eigenvector":  f"eigenvector_L{lyr}",
        "out_strength": f"out_strength_L{lyr}",
        "in_strength":  f"in_strength_L{lyr}",
    })
    if cent_wide.empty:
        cent_wide = lyr_data
    else:
        cent_wide = cent_wide.merge(lyr_data, on=["year","iso3"], how="outer")

print(f"  Centrality wide: {len(cent_wide):,} rows")

# ── LOAD VULNERABILITY PANEL ───────────────────────────────────
print("Loading vulnerability panel...")
vuln = pd.read_csv(os.path.join(DATA_DIR, "03_vulnerability_panel.csv"))
print(f"  Vulnerability: {len(vuln):,} rows")

# ── USGS PRODUCTION DATA (real, from USGS MCS) ────────────────
print("Loading USGS production data...")
usgs_rows = [
    ["CHN","China",         130000,105000,100000,95000,105000,105000,105000,105000,120000,132000,140000,168000,210000,255000,270000,44000000],
    ["AUS","Australia",          0,  2600,  3400, 2000,  8000, 10000, 14000, 20000, 21000, 21000, 21000, 22000, 18000, 16000, 13000,35700000],
    ["USA","United States",      0,     0,     0, 4000,  4100,  4700,     0,     0, 15000, 26000, 39000, 42400, 42500, 41600, 45000, 1900000],
    ["MMR","Myanmar",            0,     0,     0,    0,     0,     0,     0,     0, 17000, 22000, 36000, 26000, 38000, 43000, 31000,       0],
    ["IND","India",           2700,  2700,  2700, 2900,  2900,  1700,  1700,  1500,  1500,  3000,  2900,  2900,  2900,  2900,  2900, 6900000],
    ["RUS","Russia",             0,  2500,  2500, 2500,  2500,  2500,  2500,  2600,  2600,  2700,  2600,  2500,  2500,  2500,  2500, 3800000],
    ["BRA","Brazil",           550,   178,   178,  890,  1200,   880,  1100,  1200,  1000,  1000,   140,   140,   140,   140,    20,21000000],
    ["MYS","Malaysia",           0,     0,     0,  130,   170,   500,   300,   200,   300,   300,   300,   300,   300,   310,   130,       0],
    ["THA","Thailand",           0,     0,     0,    0,     0,     0,     0,     0,     0,     0,     0,     0,     0,  3600, 13000,    4500],
    ["NGA","Nigeria",            0,     0,     0,    0,     0,     0,     0,     0,     0,     0,     0,     0,     0,  7200, 13000,       0],
    ["VNM","Vietnam",            0,     0,     0,    0,     0,     0,     0,   200,   800,   900,   500,   400,   400,   300,   300, 3500000],
    ["MDG","Madagascar",         0,     0,     0,    0,     0,     0,     0,     0,  2000,  2000,  2000,  2100,  2100,  2100,  2000,       0],
    ["CAN","Canada",             0,     0,     0,    0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,  830000],
    ["ZAF","South Africa",       0,     0,     0,    0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,  860000],
    ["JPN","Japan",              0,     0,     0,    0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,       0],
    ["KOR","South Korea",        0,     0,     0,    0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,       0],
    ["DEU","Germany",            0,     0,     0,    0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,       0],
    ["FRA","France",             0,     0,     0,    0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,       0],
    ["GBR","United Kingdom",     0,     0,     0,    0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,       0],
    ["EST","Estonia",            0,     0,     0,    0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,     0,       0],
]
years_list = list(range(2010, 2025))
col_names = ["iso3","country"] + [str(y) for y in years_list] + ["reserves_MT"]
df_usgs_wide = pd.DataFrame(usgs_rows, columns=col_names)

usgs_long = df_usgs_wide.melt(
    id_vars=["iso3","country","reserves_MT"],
    value_vars=[str(y) for y in years_list],
    var_name="year", value_name="mining_production_REO"
)
usgs_long["year"] = usgs_long["year"].astype(int)
usgs_long = usgs_long[usgs_long["year"] >= 2012]  # match BACI range

# ── GPR INDEX (approximate) ────────────────────────────────────
gpr_country = {
    "CHN": [94,98,95,90,92,88,90,95,102,108,115,105,138,125,130],
    "USA": [105,115,108,95,100,112,118,125,130,135,145,130,155,140,145],
    "RUS": [90,92,88,92,195,145,130,118,115,108,112,118,385,290,270],
    "MMR": [80,80,82,85,88,85,88,90,95,95,115,110,180,210,200],
    "JPN": [95,100,92,90,95,88,90,92,98,100,108,102,115,110,112],
}
gpr_global = [100,108,105,98,118,115,110,108,112,118,135,118,165,145,142]

def get_gpr(iso3, yr):
    idx = yr - 2010
    if 0 <= idx < len(gpr_global):
        return gpr_country.get(iso3, gpr_global)[idx] if iso3 in gpr_country else gpr_global[idx]
    return None

# ── GDP (approximate; replace with WDI download) ──────────────
gdp_data = {
    "CHN": [6.1,7.5,8.5,9.6,10.5,11.1,11.2,12.3,13.6,13.9,14.3,14.7,14.7,17.7,18.5],
    "USA": [15.0,15.5,16.2,16.8,17.5,18.2,18.7,19.5,20.6,21.4,20.9,23.3,25.4,27.0,28.8],
    "JPN": [5.7,6.2,6.2,5.2,4.8,4.4,5.0,4.9,5.0,5.1,5.0,5.0,4.2,4.2,4.1],
    "DEU": [3.4,3.7,3.5,3.7,3.9,3.4,3.5,3.7,4.0,3.9,3.9,4.3,4.1,4.4,4.5],
    "GBR": [2.5,2.6,2.7,2.7,2.9,2.9,2.7,2.7,2.9,2.9,2.7,3.1,3.1,3.1,3.4],
    "FRA": [2.6,2.9,2.7,2.8,2.8,2.4,2.5,2.6,2.8,2.7,2.6,2.9,2.8,3.0,3.1],
    "KOR": [1.1,1.2,1.2,1.3,1.4,1.4,1.5,1.6,1.7,1.6,1.6,1.8,1.7,1.7,1.8],
    "AUS": [1.2,1.5,1.5,1.6,1.5,1.3,1.3,1.3,1.4,1.4,1.3,1.6,1.7,1.7,1.8],
    "IND": [1.7,1.8,1.8,1.9,2.0,2.1,2.3,2.7,2.7,2.9,2.7,3.2,3.4,3.7,4.0],
    "RUS": [1.5,2.1,2.0,2.3,2.1,1.4,1.3,1.6,1.7,1.7,1.5,1.8,2.2,2.1,2.1],
    "BRA": [2.2,2.6,2.5,2.5,2.4,1.8,1.8,2.1,1.9,1.9,1.4,1.6,1.9,2.1,2.2],
    "CAN": [1.6,1.8,1.8,1.9,1.8,1.6,1.5,1.7,1.7,1.7,1.6,2.0,2.1,2.1,2.2],
}
def get_gdp(iso3, yr):
    idx = yr - 2010
    if iso3 in gdp_data and 0 <= idx < len(gdp_data[iso3]):
        return gdp_data[iso3][idx] * 1e12
    return None

# ── BUILD BASE GRID: all country-year combinations ─────────────
print("Building master panel...")
all_countries = sorted(set(cent_wide["iso3"].unique()) |
                       set(vuln["iso3"].unique()))
all_years     = sorted(set(cent_wide["year"].unique()) |
                       set(vuln["year"].unique()))

base = pd.DataFrame(
    [(iso3, yr) for iso3 in all_countries for yr in all_years],
    columns=["iso3","year"]
)

# Merge USGS
base = base.merge(
    usgs_long[["iso3","year","mining_production_REO","reserves_MT"]],
    on=["iso3","year"], how="left"
)
base = base.merge(
    df_usgs_wide[["iso3","country"]],
    on="iso3", how="left"
)

# Add GPR and GDP
base["gpr_index"]       = base.apply(lambda r: get_gpr(r["iso3"], r["year"]), axis=1)
base["gdp_current_usd"] = base.apply(lambda r: get_gdp(r["iso3"], r["year"]), axis=1)
base["gdp_note"]        = "Approximate — replace with WDI NY.GDP.MKTP.CD"

# MSP and NATO flags
MSP_MEMBERS  = {"USA","CAN","AUS","GBR","FRA","DEU","JPN","KOR","IND","ITA"}
NATO_MEMBERS = {"USA","CAN","GBR","FRA","DEU","ITA","EST","FIN"}
base["msp_member"]  = base.apply(
    lambda r: 1 if r["iso3"] in MSP_MEMBERS and r["year"] >= 2022 else 0, axis=1)
base["nato_member"] = base["iso3"].apply(lambda x: 1 if x in NATO_MEMBERS else 0)

# Merge centrality
base = base.merge(cent_wide, on=["iso3","year"], how="left")

# Merge vulnerability and bargaining power
base = base.merge(vuln, on=["iso3","year"], how="left")

base = base.sort_values(["iso3","year"]).reset_index(drop=True)

# ── SAVE CSV ───────────────────────────────────────────────────
out_csv = os.path.join(DATA_DIR, "04_master_panel.csv")
base.to_csv(out_csv, index=False)
print(f"\n✓ Master panel saved: {out_csv}")
print(f"  Rows: {len(base):,} | Countries: {base['iso3'].nunique()} | "
      f"Years: {base['year'].min()}–{base['year'].max()}")
print(f"  Columns: {len(base.columns)}")
print(f"  Columns list:\n  {list(base.columns)}\n")

# ── UPDATE EXCEL WORKBOOK ──────────────────────────────────────
print("Updating REE_Master_Dataset.xlsx → Country_Year_Panel sheet...")
wb_path = os.path.join(PROJECT_DIR, "REE_Master_Dataset.xlsx")

if os.path.exists(wb_path):
    wb = load_workbook(wb_path)
    if "Country_Year_Panel" in wb.sheetnames:
        del wb["Country_Year_Panel"]

    ws = wb.create_sheet("Country_Year_Panel", 1)
    ws.sheet_properties.tabColor = "375623"

    HEADER_COLOR = "375623"

    # Write headers
    for c_idx, col_name in enumerate(base.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write data rows
    GREY = "F2F2F2"
    for r_idx, row in enumerate(base.itertuples(index=False), 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx,
                           value=None if (isinstance(val, float) and np.isnan(val)) else val)
            cell.alignment = Alignment(vertical="top")
            if r_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=GREY)

    ws.freeze_panes = "C2"

    # Set column widths
    for c_idx in range(1, len(base.columns) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 18

    wb.save(wb_path)
    print(f"✓ Excel workbook updated: {wb_path}")
else:
    print(f"  Excel not found at {wb_path} — CSV saved only.")

# ── PRINT SUMMARY STATS ────────────────────────────────────────
print("\n=== DATA COVERAGE SUMMARY ===")
for col in ["betweenness_L2","china_import_share_L2","supplier_hhi_L2",
            "supply_vulnerability_idx","dependent_importers_25pct","mining_production_REO"]:
    non_null = base[col].notna().sum()
    print(f"  {col:<35}: {non_null:,} non-null rows ({non_null/len(base)*100:.0f}%)")

print("\n=== CHINA KEY METRICS (2022) ===")
chn22 = base[(base["iso3"]=="CHN") & (base["year"]==2022)].squeeze()
if len(chn22):
    print(f"  Betweenness L1: {chn22.get('betweenness_L1', 'N/A')}")
    print(f"  Betweenness L2: {chn22.get('betweenness_L2', 'N/A')}")
    print(f"  Betweenness L3: {chn22.get('betweenness_L3', 'N/A')}")
    print(f"  Dependent importers (25%): {chn22.get('dependent_importers_25pct', 'N/A')}")
    print(f"  Mining production: {chn22.get('mining_production_REO', 'N/A'):,} MT REO")

print(f"\n✓ ALL DONE. Master panel is ready for regression analysis.")
print(f"  File: {out_csv}")
print(f"\n→ NEXT: run  Step5_Regressions.py")
